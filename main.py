import requests
import pandas as pd
from openpyxl.styles import PatternFill, Font
import openpyxl as xl
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import slicing as sl
import genre_id


api_key = "b88e2294a28b94af81e9bfe6d10deee3"
url=f"https://api.themoviedb.org/3/discover/"
print("1.Movies\n2.TV Shows\n3.Help\n")
user=input("Enter(1-4): ")
position=42
while True:
    if user=='1':
        url+="".join(f"movie?api_key={api_key}")
        break
    elif user=='2':
        url+="".join(f"tv?api_key={api_key}")
        position = 40
        break
    elif user=='3':
        print("Type 1 for Movies\nType 2 for TV Shows\n")
    else:
        print("Invalid")
print("1.Search by name\n2.Search by genre\n3.Search by release date\n4.Search by rating\n5.Help")
movie_search=input("Enter(1-3): ")
while True:
    if movie_search=='1':
        name=input("Enter:")
        url+="".join(f"/{sl.string_slicing(url,name,position)}")
        break
    elif movie_search=='2':
        genre=sl.genre_entry()
        genre=genre_id.key_finder(genre)
        url=sl.add_to_url(url,genre)
        break
    elif movie_search=="3":
        release_date_gte=input("Enter Starting date(YYYY-MM-DD): ")
        release_date_lte=input("Enter Ending date(YYYY-MM-DD): ")
        url+="".join(f"&primary_release_date_gte={release_date_gte}&primary_release_date_lte={release_date_lte}")
        break
    elif movie_search=="4":
        rating_gte=input("Enter Starting rating: ")
        rating_lte=input("Enter Starting rating: ")
        url+="".join(f"&sort_by=vote_average_gte={rating_gte}&vote_average_lte={rating_lte}")
        break
    elif movie_search=='4':
        print("Type 1 to seach by name\nType 2 to seach by genre\nType 3 to seach by release date\nType 4 to seach by rating\n")
    else:
        print("INVALID")
req=requests.get(url)
if req.status_code==200:
    data=req.json()
    result=data.get("results",[])
    if results:
        df=pd.DataFrame(result)
        poped_items=["overview","backdrop_path","video","adult"]
        for item in poped_items:
            if item in df:
                df.pop(item)
        user_sort = input("Sort By (title, vote_average, vote_count): ").lower()
        user_sort_order = input("Sort in (D)escending or (A)scending order: ").lower()
        ascending = True if user_sort_order == "a" else False
        sort_options = {
            "title": "title",
            "vote_average": "vote_average",
            "vote_count": "vote_count"
        }
        pdf=df
        if user_sort in sort_options and user_sort_order in ["a", "d"] and user_sort in df:
            pdf = df.sort_values(by=sort_options[user_sort], ascending=ascending)
        else:
            print("Invalid input. Please enter valid sort option and order.")
        excel_file_path = "film_data.xlsx"
        pdf.to_excel(excel_file_path, index=False)
        wb = xl.load_workbook(excel_file_path)
        ws = wb.active
        fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        font_white = Font(color="FFFFFF")
        cells_to_style = ["A1", "C1", "E1", "G1", "I1"]
        for cell in cells_to_style:
            ws[cell].fill = fill_black
            ws[cell].font = font_white
        plt.figure(figsize=(8,6))
        plt.bar(df["original_title"], df["vote_average"], color='royalblue')
        plt.title("Vote Average by Title")
        plt.xlabel("Title")
        plt.ylabel("Vote Average")
        image_path="chart.png"
        plt.savefig(image_path)
        plt.show()
        plt.close()
        chart=Image(image_path)
        ws.add_image(chart,"N1")
        wb.save(excel_file_path)
    else:
        print("No results found")
else:
    print("API request failed with status code: ",req.status_code)